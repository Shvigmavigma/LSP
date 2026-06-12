from io import BytesIO
from pathlib import Path
import re
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


IMPORT_SPECS: Dict[str, Dict[str, Any]] = {
    "student_emails": {"columns": ["email"], "example": ["student@lit1533.ru"]},
    "teacher_emails": {"columns": ["email"], "example": ["teacher@example.ru"]},
    "students": {
        "columns": ["fullname", "email", "password", "class"],
        "example": ["Иванов Иван Иванович", "student@lit1533.ru", "StrongPassword123", 10.1],
    },
    "teachers": {
        "columns": ["fullname", "email", "password"],
        "example": ["Петров Петр Петрович", "teacher@example.ru", "StrongPassword123"],
    },
    "projects": {
        "columns": ["title", "body", "customer_email"],
        "example": ["Школьный проект", "Краткое описание проекта", "customer@example.ru"],
    },
}
EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
CUSTOM_TEMPLATES_DIR = Path("excel_templates")


class ExcelImportValidationError(ValueError):
    pass


def create_excel_template(
    import_type: str,
    example_values: Dict[str, Any] | None = None,
    columns: List[str] | None = None,
) -> BytesIO:
    spec = IMPORT_SPECS.get(import_type)
    if not spec:
        raise ExcelImportValidationError("Неизвестный тип импорта")

    template_columns = columns or list(spec["columns"])
    validate_template_columns(import_type, template_columns)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "import"
    worksheet.append(template_columns)
    example = [
        (example_values or {}).get(
            column,
            spec["example"][spec["columns"].index(column)] if column in spec["columns"] else "",
        )
        for column in template_columns
    ]
    worksheet.append(example)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F855A")
    for index, column in enumerate(template_columns, start=1):
        sample = str(example[index - 1])
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = max(
            len(column) + 4, len(sample) + 4
        )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def get_excel_template(import_type: str) -> BytesIO:
    if import_type not in IMPORT_SPECS:
        raise ExcelImportValidationError("Неизвестный тип импорта")
    custom_template = CUSTOM_TEMPLATES_DIR / f"{import_type}.xlsx"
    if custom_template.exists():
        return BytesIO(custom_template.read_bytes())
    return create_excel_template(import_type)


def get_excel_template_config(import_type: str) -> Dict[str, Any]:
    spec = IMPORT_SPECS.get(import_type)
    if not spec:
        raise ExcelImportValidationError("Неизвестный тип импорта")
    custom_template = CUSTOM_TEMPLATES_DIR / f"{import_type}.xlsx"
    template_columns = list(spec["columns"])
    examples = dict(zip(template_columns, spec["example"]))
    if custom_template.exists():
        workbook = load_workbook(custom_template, read_only=True, data_only=True)
        try:
            rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True, max_row=2))
            if rows:
                template_columns = [str(value).strip() for value in rows[0] if value is not None]
            if len(rows) > 1:
                examples = {
                    column: rows[1][index] if index < len(rows[1]) and rows[1][index] is not None else ""
                    for index, column in enumerate(template_columns)
                }
        finally:
            workbook.close()
    return {
        "columns": template_columns,
        "required_columns": list(spec["columns"]),
        "example_values": examples,
        "is_custom": custom_template.exists(),
    }


def install_excel_template(content: bytes, import_type: str) -> None:
    # A custom template may contain examples and formatting, but its required
    # header must remain identical to the system format.
    parse_excel_headers(content, import_type, use_current_template=False)
    CUSTOM_TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    (CUSTOM_TEMPLATES_DIR / f"{import_type}.xlsx").write_bytes(content)


def generate_excel_template(
    import_type: str,
    example_values: Dict[str, Any],
    columns: List[str],
) -> BytesIO:
    spec = IMPORT_SPECS.get(import_type)
    if not spec:
        raise ExcelImportValidationError("Неизвестный тип импорта")
    validate_template_columns(import_type, columns)
    if set(example_values) != set(columns):
        raise ExcelImportValidationError("Для каждого столбца должен быть указан пример")
    if any(value is None or str(value).strip() == "" for value in example_values.values()):
        raise ExcelImportValidationError("Заполните примеры для всех обязательных колонок")
    stream = create_excel_template(import_type, example_values, columns)
    install_excel_template(stream.getvalue(), import_type)
    stream.seek(0)
    return stream


def reset_excel_template(import_type: str) -> bool:
    if import_type not in IMPORT_SPECS:
        raise ExcelImportValidationError("Неизвестный тип импорта")
    custom_template = CUSTOM_TEMPLATES_DIR / f"{import_type}.xlsx"
    if custom_template.exists():
        custom_template.unlink()
        return True
    return False


def validate_template_columns(import_type: str, columns: List[str]) -> None:
    spec = IMPORT_SPECS.get(import_type)
    if not spec:
        raise ExcelImportValidationError("Неизвестный тип импорта")
    normalized = [str(column).strip() for column in columns]
    if any(not column for column in normalized):
        raise ExcelImportValidationError("Названия столбцов не могут быть пустыми")
    if len(set(normalized)) != len(normalized):
        raise ExcelImportValidationError("Названия столбцов не должны повторяться")
    missing = [column for column in spec["columns"] if column not in normalized]
    if missing:
        raise ExcelImportValidationError(
            f"Нельзя удалить обязательные столбцы: {', '.join(missing)}"
        )


def get_expected_columns(import_type: str) -> List[str]:
    return get_excel_template_config(import_type)["columns"]


def parse_excel_headers(
    content: bytes,
    import_type: str,
    use_current_template: bool = True,
) -> List[str]:
    spec = IMPORT_SPECS.get(import_type)
    if not spec:
        raise ExcelImportValidationError("Неизвестный тип импорта")
    if not content:
        raise ExcelImportValidationError("Файл пуст")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelImportValidationError("Не удалось прочитать XLSX-файл") from exc
    if len(workbook.sheetnames) != 1:
        raise ExcelImportValidationError("В файле должен быть ровно один лист")
    worksheet = workbook[workbook.sheetnames[0]]
    first_row = next(worksheet.iter_rows(values_only=True), None)
    if not first_row:
        raise ExcelImportValidationError("В файле отсутствует строка заголовков")
    headers = [str(value).strip() if value is not None else "" for value in first_row]
    expected = get_expected_columns(import_type) if use_current_template else None
    if expected is None:
        validate_template_columns(import_type, headers)
        return headers
    if headers != expected:
        raise ExcelImportValidationError(
            f"Колонки должны строго соответствовать шаблону: {', '.join(expected)}"
        )
    return headers


def parse_excel_import(content: bytes, import_type: str) -> List[Dict[str, Any]]:
    spec = IMPORT_SPECS.get(import_type)
    if not spec:
        raise ExcelImportValidationError("Неизвестный тип импорта")
    expected = parse_excel_headers(content, import_type)
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelImportValidationError("Не удалось прочитать XLSX-файл") from exc
    rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))
    required = spec["columns"]

    parsed: List[Dict[str, Any]] = []
    seen_emails = set()
    for excel_row, values in enumerate(rows[1:], start=2):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        if len(values) != len(expected):
            raise ExcelImportValidationError(f"Строка {excel_row}: неверное количество колонок")
        item = {
            column: value.strip() if isinstance(value, str) else value
            for column, value in zip(expected, values)
        }
        missing = [
            column for column in required
            if item[column] is None or str(item[column]).strip() == ""
        ]
        if missing:
            raise ExcelImportValidationError(
                f"Строка {excel_row}: не заполнены обязательные поля: {', '.join(missing)}"
            )
        email_field = "customer_email" if import_type == "projects" else "email"
        email = str(item[email_field]).strip().lower()
        if not EMAIL_PATTERN.fullmatch(email):
            raise ExcelImportValidationError(f"Строка {excel_row}: некорректная почта")
        if import_type != "projects" and email in seen_emails:
            raise ExcelImportValidationError(f"Строка {excel_row}: почта повторяется в файле")
        seen_emails.add(email)
        item[email_field] = email
        if import_type == "students":
            try:
                item["class"] = float(item["class"])
            except (TypeError, ValueError) as exc:
                raise ExcelImportValidationError(
                    f"Строка {excel_row}: класс должен быть числом, например 10.1"
                ) from exc
        item["_row"] = excel_row
        parsed.append(item)
    if not parsed:
        raise ExcelImportValidationError("После заголовка нет данных для импорта")
    return parsed
